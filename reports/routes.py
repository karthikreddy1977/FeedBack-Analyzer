"""
Pulse — Reports Routes
========================
Generate downloadable reports in CSV, Excel (openpyxl), and PDF (fpdf2).
Reports include feedback, sentiment, emotion, category, scores, and analytics.
"""

import os
import csv
import io
from datetime import datetime, timezone
from collections import Counter

from flask import (
    render_template, request, jsonify, send_file,
    redirect, url_for, flash, current_app
)
from flask_login import login_required, current_user

from . import reports_bp
from ..extensions import db
from ..models import Feedback, Report
from ..utils import create_notification


HEADERS = ["ID", "Feedback", "Category", "Sentiment", "Polarity", "Subjectivity", "Emotion", "Keywords", "Date"]


# ---------------------------------------------------------------------------
# Reports List
# ---------------------------------------------------------------------------
@reports_bp.route("/")
@login_required
def reports_list():
    """Display the user's generated reports."""
    reports = Report.query.filter_by(user_id=current_user.id) \
        .order_by(Report.created_at.desc()).all()
    return render_template("reports/list.html", reports=reports)


# ---------------------------------------------------------------------------
# Generate Report
# ---------------------------------------------------------------------------
@reports_bp.route("/generate", methods=["POST"])
@login_required
def generate_report():
    """Generate a report in the requested format (csv, xlsx, pdf)."""
    fmt = request.form.get("format", "csv").lower()
    if fmt not in ("csv", "xlsx", "pdf"):
        flash("Invalid report format.", "error")
        return redirect(url_for("reports.reports_list"))

    feedbacks = Feedback.query.filter_by(user_id=current_user.id) \
        .order_by(Feedback.created_at.desc()).all()

    if not feedbacks:
        flash("No feedback data to export.", "error")
        return redirect(url_for("reports.reports_list"))

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    export_dir = current_app.config.get("EXPORT_DIR", "exports")
    os.makedirs(export_dir, exist_ok=True)

    # Build data rows
    rows = []
    for fb in feedbacks:
        rows.append([
            fb.id,
            fb.feedback_text,
            fb.category,
            fb.sentiment,
            f"{fb.polarity_score:.4f}",
            f"{fb.subjectivity_score:.4f}",
            fb.emotion,
            fb.keywords,
            fb.created_at.strftime("%Y-%m-%d %H:%M") if fb.created_at else "",
        ])

    # Analytics summary
    total = len(feedbacks)
    sentiments = Counter(fb.sentiment for fb in feedbacks)
    avg_polarity = sum(fb.polarity_score for fb in feedbacks) / total
    avg_subj = sum(fb.subjectivity_score for fb in feedbacks) / total

    try:
        if fmt == "csv":
            filename, filepath = _generate_csv(rows, timestamp, export_dir)
        elif fmt == "xlsx":
            filename, filepath = _generate_xlsx(rows, sentiments, avg_polarity, avg_subj, total, timestamp, export_dir)
        elif fmt == "pdf":
            filename, filepath = _generate_pdf(rows, sentiments, avg_polarity, avg_subj, total, timestamp, export_dir)

        # Save report record
        report = Report(
            user_id=current_user.id,
            report_name=filename,
            file_path=filepath,
            format=fmt,
        )
        db.session.add(report)
        db.session.commit()

        create_notification(current_user.id, f"Report generated: {filename} 📄")
        flash(f"Report generated successfully: {filename}", "success")

    except Exception as e:
        flash(f"Error generating report: {str(e)}", "error")

    return redirect(url_for("reports.reports_list"))


def _generate_csv(rows, timestamp, export_dir):
    """Generate a CSV report."""
    filename = f"pulse_report_{timestamp}.csv"
    filepath = os.path.join(export_dir, filename)

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        writer.writerows(rows)

    return filename, filepath


def _generate_xlsx(rows, sentiments, avg_polarity, avg_subj, total, timestamp, export_dir):
    """Generate an Excel report with openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    filename = f"pulse_report_{timestamp}.xlsx"
    filepath = os.path.join(export_dir, filename)

    wb = Workbook()

    # --- Data Sheet ---
    ws = wb.active
    ws.title = "Feedback Data"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="16213D", end_color="16213D", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Auto-width columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    # --- Summary Sheet ---
    ws2 = wb.create_sheet("Analytics Summary")
    summary_data = [
        ["Metric", "Value"],
        ["Total Feedback", total],
        ["Positive", sentiments.get("Positive", 0)],
        ["Negative", sentiments.get("Negative", 0)],
        ["Neutral", sentiments.get("Neutral", 0)],
        ["Average Polarity", f"{avg_polarity:.4f}"],
        ["Average Subjectivity", f"{avg_subj:.4f}"],
    ]
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.border = thin_border

    wb.save(filepath)
    return filename, filepath


def _generate_pdf(rows, sentiments, avg_polarity, avg_subj, total, timestamp, export_dir):
    """Generate a PDF report with fpdf2."""
    from fpdf import FPDF

    filename = f"pulse_report_{timestamp}.pdf"
    filepath = os.path.join(export_dir, filename)

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Title
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 12, "Pulse - Feedback Analysis Report", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 8, f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(10)

    # Analytics Summary
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Analytics Summary", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 7, f"Total Feedback: {total}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 7, f"Positive: {sentiments.get('Positive', 0)}  |  Negative: {sentiments.get('Negative', 0)}  |  Neutral: {sentiments.get('Neutral', 0)}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 7, f"Average Polarity: {avg_polarity:.4f}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 7, f"Average Subjectivity: {avg_subj:.4f}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(10)

    # Feedback Table
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Feedback Details", new_x="LMARGIN", new_y="NEXT")

    # Table header
    pdf.set_font("Helvetica", "B", 8)
    col_widths = [12, 60, 22, 22, 18, 18, 20, 18]
    headers_short = ["ID", "Feedback", "Category", "Sentiment", "Polarity", "Subj.", "Emotion", "Date"]

    pdf.set_fill_color(22, 33, 61)
    pdf.set_text_color(255, 255, 255)
    for i, header in enumerate(headers_short):
        pdf.cell(col_widths[i], 7, header, border=1, fill=True, align="C")
    pdf.ln()

    # Table rows
    pdf.set_font("Helvetica", "", 7)
    pdf.set_text_color(0, 0, 0)
    for row in rows[:200]:  # Limit to 200 rows in PDF for readability
        max_h = 7
        # Truncate feedback text for PDF
        feedback_text = str(row[1])[:80] + ("..." if len(str(row[1])) > 80 else "")
        row_data = [str(row[0]), feedback_text, str(row[2]), str(row[3]),
                    str(row[4]), str(row[5]), str(row[6]), str(row[8])]

        for i, val in enumerate(row_data):
            pdf.cell(col_widths[i], max_h, val, border=1, align="C" if i != 1 else "L")
        pdf.ln()

    pdf.output(filepath)
    return filename, filepath


# ---------------------------------------------------------------------------
# Download Report
# ---------------------------------------------------------------------------
@reports_bp.route("/download/<int:report_id>")
@login_required
def download_report(report_id):
    """Download a generated report (owner only)."""
    report = Report.query.filter_by(
        id=report_id, user_id=current_user.id
    ).first_or_404()

    if not os.path.exists(report.file_path):
        flash("Report file not found.", "error")
        return redirect(url_for("reports.reports_list"))

    mime_types = {
        "csv": "text/csv",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "pdf": "application/pdf",
    }

    return send_file(
        report.file_path,
        mimetype=mime_types.get(report.format, "application/octet-stream"),
        as_attachment=True,
        download_name=report.report_name,
    )


# ---------------------------------------------------------------------------
# Delete Report
# ---------------------------------------------------------------------------
@reports_bp.route("/<int:report_id>/delete", methods=["POST"])
@login_required
def delete_report(report_id):
    """Delete a report and its file (owner only)."""
    report = Report.query.filter_by(
        id=report_id, user_id=current_user.id
    ).first_or_404()

    # Remove file from disk
    if os.path.exists(report.file_path):
        os.remove(report.file_path)

    db.session.delete(report)
    db.session.commit()

    flash("Report deleted.", "success")
    return redirect(url_for("reports.reports_list"))
